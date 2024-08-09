import pandas as pd
import openpyxl as xl
from openpyxl.utils.dataframe import dataframe_to_rows
import datetime as dt
import os
try:
    from db_engine_connector import create_mav_engine
except:
    from python_scripts.db_engine_connector import create_mav_engine


def create_se_report():
    engine=create_mav_engine()
    
    pta_call="""With cte as 
                    (select EntityId, 
                        max(CreatedAt) as CreatedAt 
                        from MDR.dbo.ChangeAudits 
                    where 
                        FieldName_Formatted='Status' 
                        and newFieldValue = 'Planning' group by EntityId)



                Select 
                    CONCAT('=HYPERLINK("https://app.medicaiddoneright.com/application-page/',app.[Id],'", "', APP.FileNumber ,'")')  AS [File #],
                    Processor_FullName as Processor,
                    FullName as Resident,
                    Facility_Name as Facility,
                    Facility_County as County,
                    Facility_State_AsString as State,
                    ProcessingAge,
                    datediff(D,ProcessorComments_LastCreatedDate,GETDATE()) as [Last Processor Comment Age],
                    datediff(D,audit.CreatedAt,GetDate()) - (datediff(wk,audit.CreatedAt,GetDate())*2)
                            -Case when DATEPART(dw,audit.CreatedAt)=1 Then 1 else 0 end
                            + case when datepart(dw,GetDate())=1 Then 1 else 0 end as [Time in Status]



                from MDR.dbo.Application as app
                    inner join cte as audit on audit.Entityid= app.Id

                where 
                    [ProductType_AsString] in ('New Application', 'Conversion to LTC', 'Guardianship')
                    and InternalStatus_AsString = 'Open'
                    and Status_AsString ='Planning'

                    and app.Facility_State_AsString in ('GA','NC','VA')
                    and datediff(D,audit.CreatedAt,GetDate()) - (datediff(wk,audit.CreatedAt,GetDate())*2)
                            -Case when DATEPART(dw,audit.CreatedAt)=1 Then 1 else 0 end
                            + case when datepart(dw,GetDate())=1 Then 1 else 0 end >=2
                order by [Time in Status] desc"""


    Ats_call="""With cte as 
                (select EntityId,max(CreatedAt) as CreatedAt from MDR.dbo.ChangeAudits where FieldName_Formatted='Status' and newFieldValue = 'Applied' group by EntityId)



                Select 
                CONCAT('=HYPERLINK("https://app.medicaiddoneright.com/application-page/',app.[Id],'", "', APP.FileNumber ,'")')  AS [File #],
                Processor_FullName as Processor,
                FullName as Resident,
                Facility_Name as Facility,
                Facility_County as County,
                Facility_State_AsString as State,
                ProcessingAge,
                datediff(D,ProcessorComments_LastCreatedDate,GETDATE()) as [Last Processor Comment Age],
                datediff(D,audit.CreatedAt,GetDate()) as [Time in Status]


                from MDR.dbo.Application as app
                 inner join cte as audit on audit.Entityid= app.Id

                where [ProductType_AsString] in ('New Application', 'Conversion to LTC', 'Guardianship')
                and InternalStatus_AsString = 'Open'
                and Status_AsString ='Applied'
         
                and app.Facility_State_AsString in ('GA','NC','VA')
                and datediff(D,audit.CreatedAt,GetDate()) >=45

                order by [Time in Status] desc"""
    
    SF_call= """With cte as 
                (select EntityId, max(CreatedAt) as CreatedAt from MDR.dbo.ChangeAudits where FieldName_Formatted='Status' and newFieldValue = 'Submitted' group by EntityId)



                Select 
                CONCAT('=HYPERLINK("https://app.medicaiddoneright.com/application-page/',app.[Id],'", "', APP.FileNumber ,'")')  AS [File #],
                Processor_FullName as Processor,
                FullName as Resident,
                Facility_Name as Facility,
                Facility_County as County,
                Facility_State_AsString as State,
                ProcessingAge,
                datediff(D,ProcessorComments_LastCreatedDate,GETDATE()) as [Last Processor Comment Age],
                datediff(D,audit.CreatedAt,GetDate()) as [Time in Status]


                from MDR.dbo.Application as app
                inner join cte as audit on audit.Entityid= app.Id

                where [ProductType_AsString] in ('New Application', 'Conversion to LTC', 'Guardianship')
                and InternalStatus_AsString = 'Open'
                and Status_AsString ='Submitted'
               
                and app.Facility_State_AsString in ('GA','NC','VA')
                and datediff(D,audit.CreatedAt,GetDate()) >=7

                order by [Time in Status] desc"""
    
    pta_db=pd.read_sql(pta_call,engine)
    ats_db=pd.read_sql(Ats_call,engine)
    sf_db=pd.read_sql(SF_call,engine)

    sheets=['Planning to Applied','Applied to Submitted','Submitted Files']
    dataframes=[pta_db,ats_db,sf_db]

    workbook=xl.Workbook()

    for sheet, df in zip(sheets, dataframes):
        ws = workbook.create_sheet(title=sheet)
        for row in dataframe_to_rows(df, index=False, header=True):
            ws.append(row)
    
    
    #wks formatting
    workbook.remove(workbook['Sheet'])

    def as_text(value):
        if value is None:
            return ""
        return str(value)

    for ws in workbook:
        for column_cells in ws.columns:
            if column_cells[0].column==1:
                ws.column_dimensions[xl.utils.get_column_letter(column_cells[0].column)].width = 14
            else:
                length = max(len(as_text(cell.value)) for cell in column_cells) + 2
                ws.column_dimensions[xl.utils.get_column_letter(column_cells[0].column)].width = length


    date=dt.datetime.now().strftime('%m-%d-%y')
    if os.name == 'nt':
        file_path=f"outputs/Status Exceptions Report ({date}).xlsx"
    else:
        file_path= os.path.join(os.getcwd(), "dags", "python_scripts", "Status_Exceptions", "outputs",f"Status Exceptions Report ({date}).xlsx")
    
    workbook.save(file_path)
    workbook.close()

